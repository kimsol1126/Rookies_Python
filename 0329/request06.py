from bs4 import BeautifulSoup
import requests
from datetime import datetime 
import openpyxl
import time
import schedule

def boannews():
    now = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet['A1'] = "제목"
    worksheet['B1'] = "내용"

    url = "http://www.boannews.com/media/t_list.asp"

    headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
    }

    req = requests.get(url, headers=headers)
    soup = BeautifulSoup(req.text, "lxml")

    boan_news = soup.select("#news_area > div > a > span")
    news_content = soup.select("#news_area > div > a.news_content")

    row = 2
    for news, content in zip(boan_news, news_content):
        print(f"제목: {news.string} \n내용: {content.string}")
        worksheet.cell(row=row, column=1, value=news.string)
        worksheet.cell(row=row, column=2, value=content.string)
        row = row + 1

    workbook.save(f"{now}.xlsx")


schedule.every(1).days.do(boannews) #3분마다 job 실행

while True:
    schedule.run_pending()
    time.sleep(1)