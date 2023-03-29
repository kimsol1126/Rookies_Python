import openpyxl
from faker import Faker

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = "Name"
worksheet['B1'] = "Email"
worksheet['C1'] = "Phone"

fake = Faker('ko_KR') #한국 데이터 생성

for row in range(2, 50):
    worksheet.cell(row=row, column=1, value=fake.name())
    worksheet.cell(row=row, column=2, value=fake.email())
    worksheet.cell(row=row, column=3, value=fake.phone_number())

workbook.save('member.xlsx')